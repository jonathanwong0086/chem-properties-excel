# -*- coding: utf-8 -*-
"""
化学品物性数据Excel生成模板
用法: 修改 chemicals 列表后运行此脚本
"""
import openpyxl, os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment

# ========== 在此处填入化学品数据 ==========
# 每个化学品是一个 dict，包含：
#   name, cas_no, sources (每个字段一个来源URL),
#   mp, bp, fp, density, state, fire, sol, den_water, water_ext, explosion, vapor, oel, danger_cat

chemicals = [
    # 示例：
    # {
    #     "name": "乙醇", "cas_no": "64-17-5",
    #     "sources": {
    #         "cas": "PubChem (https://pubchem.ncbi.nlm.nih.gov/compound/702)",
    #         "mp": "NIST Chemistry WebBook (https://webbook.nist.gov/cgi/cbook.cgi?ID=64-17-5)",
    #         "bp": "ChemicalBook (https://www.chemicalbook.com/)",
    #         "fp": "Sigma-Aldrich SDS (https://www.sigmaaldrich.com/)",
    #         "density": "PubChem (https://pubchem.ncbi.nlm.nih.gov/compound/702)",
    #         "state": "ChemSpider (http://www.chemspider.com/)",
    #         "fire": "GB 50016-2014《建筑设计防火规范》表3.1.1",
    #         "sol": "PubChem",
    #         "den_water": "计算值: 0.789/1.0",
    #         "water_ext": "CAMEO Chemicals NOAA (https://cameochemicals.noaa.gov/)",
    #         "explosion": "NIOSH Pocket Guide (https://www.cdc.gov/niosh/npg/npgd0188.html)",
    #         "vapor": "NIST Chemistry WebBook (MR=1.59)",
    #         "oel": "GBZ 2.1-2019 表1 (https://openstd.samr.gov.cn/)",
    #         "danger_cat": "《危险化学品目录》(2015版) 第2568项",
    #     },
    #     "mp": "-114.1", "bp": "78.3", "fp": "12",
    #     "density": "0.789", "state": "液体",
    #     "fire": "甲类", "sol": "与水混溶",
    #     "den_water": "否", "water_ext": "否",
    #     "explosion": "3.3%~19%", "vapor": "是 (MR=1.59)",
    #     "oel": "TWA 200", "danger_cat": "易燃液体 + 有毒",
    # },
]

# ========== 输出配置 ==========
PROJECT_NAME = "K33工艺流程"  # 修改为实际项目名
OUTPUT_DIR = r"C:\Users\Jonathan\Desktop"

# ========== 数据来源汇总 ==========
SRC_DATA = [
    ("国际数据库", "PubChem (NCBI)", "CAS号、密度、水溶性", "https://pubchem.ncbi.nlm.nih.gov/"),
    ("国际数据库", "NIST Chemistry WebBook", "熔点、沸点、蒸气相对密度", "https://webbook.nist.gov/"),
    ("化学品供应商", "Sigma-Aldrich / MilliporeSigma", "安全技术说明书(SDS)、闪点、密度", "https://www.sigmaaldrich.com/"),
    ("化学品数据库", "ChemicalBook", "CAS号、沸点、闪点、物态、水溶性", "https://www.chemicalbook.com/"),
    ("化学品数据库", "ChemSpider (Royal Society of Chemistry)", "常温物态、结构验证", "http://www.chemspider.com/"),
    ("应急响应数据库", "CAMEO Chemicals (NOAA)", "遇水反应性、灭火禁忌、爆炸极限", "https://cameochemicals.noaa.gov/"),
    ("职业安全数据库", "NIOSH Pocket Guide (CDC)", "爆炸极限(LEL/UEL)、蒸气密度", "https://www.cdc.gov/niosh/npg/"),
    ("中国国家标准", "GB 50016-2014《建筑设计防火规范》", "火灾危险性类别 (甲/乙/丙类)", "http://www.mohurd.gov.cn/"),
    ("中国国家标准", "GBZ 2.1-2019《工作场所有害因素职业接触限值》", "职业接触限值 OEL (TWA/STEL/MAC)", "https://openstd.samr.gov.cn/"),
    ("中国国家标准", "GB 15603-1995《常用危险化学品贮存通则》", "储存禁忌、遇湿燃烧品判定", "https://openstd.samr.gov.cn/"),
    ("中国国家标准", "GB/T 42594-2023《承压设备介质危害分类导则》", "毒性分级 (I/II级) 辅助参考", "https://openstd.samr.gov.cn/"),
    ("中国部门规章", "《危险化学品目录》(2015版) — 应急管理部", "危化品类别标注 (易燃液体/腐蚀品/氧化性等)", "https://www.mem.gov.cn/"),
]

# ========== 样式定义 ==========
FIELDS = [
    ("序号","id"),("化学品名称","name"),("CAS号","cas_no"),
    ("熔点(°C)","mp"),("沸点(°C)","bp"),("闪点(°C)","fp"),
    ("密度","density"),("常温物态","state"),("火灾危险性类别","fire"),
    ("水溶性","sol"),("密度>水？","den_water"),("禁用水灭火？","water_ext"),
    ("爆炸极限(LEL~UEL)","explosion"),("蒸汽>空气？","vapor"),
    ("职业接触限值(mg/m³)","oel"),("危化品类别","danger_cat"),
]
COLOR_COLS = {2, 3, 9}  # 名称、CAS、火灾类别着色

def build_excel(chemicals, project_name, output_dir, src_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{project_name}化学品物性数据"

    # 样式
    hf = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    df = Font(name="微软雅黑", size=9)
    ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al = Alignment(horizontal="left", vertical="center", wrap_text=True)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    org = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    yel = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    wht = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # 表头
    for ci, (h, _) in enumerate(FIELDS, 1):
        c = ws.cell(1, ci, h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
    ws.row_dimensions[1].height = 32

    # 数据行
    for ri, chem in enumerate(chemicals, 2):
        fv = str(chem.get("fire", ""))
        if fv.startswith("甲"):   rfill = red
        elif fv.startswith("乙"): rfill = org
        elif fv.startswith("丙"): rfill = yel
        else:                     rfill = wht

        for ci, (_, fk) in enumerate(FIELDS, 1):
            val = (ri - 1) if fk == "id" else str(chem.get(fk, ""))
            c = ws.cell(ri, ci, val)
            c.font = df; c.border = tb
            c.alignment = al if fk in ("name","sol","explosion","oel","danger_cat","water_ext") else ac
            c.fill = rfill if ci in COLOR_COLS else wht

            src = chem.get("sources", {}).get(fk, "")
            if src and fk not in ("id", "name"):
                c.comment = Comment(f"数据来源:\n{src}", "物性查询")
                c.comment.width = 350
                c.comment.height = 70

    # 列宽
    widths = [5, 26, 20, 20, 20, 22, 22, 14, 20, 22, 10, 22, 22, 14, 22, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    for ri in range(2, len(chemicals) + 2):
        ws.row_dimensions[ri].height = 30

    # 图例
    lr = len(chemicals) + 3
    ws.cell(lr, 1, "图例:").font = Font(name="微软雅黑", bold=True, size=9)
    for r, t, f in [
        (lr+1, "甲类 = 闪点<28°C液体 / 爆炸下限<10%气体", red),
        (lr+2, "乙类 = 28°C≤闪点<60°C液体", org),
        (lr+3, "丙类 = 闪点≥60°C液体 / 可燃固体", yel),
        (lr+4, "白色 = 非可燃物质", wht),
    ]:
        ws.cell(r, 1, "").fill = f; ws.cell(r, 1).border = tb
        ws.cell(r, 2, t).font = Font(name="微软雅黑", size=9)

    # 说明
    nr = lr + 6
    notes = [
        "说明:",
        "1. 每个单元格右上角红三角 = 来源注释，同一行不同列来源不同。",
        "2. 火灾危险性类别仅供参考，实际分类应以当地消防部门核定为准。",
        "3. 仅标注甲/乙/丙类，非可燃物质统一标注为\"非可燃\"。",
        "4. 职业接触限值依据 GBZ 2.1-2019，单位 mg/m³。",
        "5. 危化品类别依据《危险化学品目录》(2015版)。",
        "6. 因牌号/结构未定的物质，标注\"因牌号而异\"，需查阅实际供应商SDS。",
    ]
    for i, n in enumerate(notes):
        ws.cell(nr + i, 1, n).font = Font(name="微软雅黑", bold=(i == 0), size=9)

    # ========== Sheet 2: 数据来源汇总 ==========
    sw = wb.create_sheet("数据来源汇总", -1)
    sw.merge_cells("A1:D1")
    for c in range(1, 5):
        sw.cell(1, c).fill = hfill; sw.cell(1, c).border = tb
    sw.cell(1, 1, f"{project_name}化学品物性数据 — 数据来源汇总").font = Font(
        name="微软雅黑", bold=True, size=12, color="FFFFFF")
    sw.cell(1, 1).fill = hfill; sw.cell(1, 1).alignment = ha
    sw.row_dimensions[1].height = 32

    shfill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for ci, h in enumerate(["类型", "机构 / 网站名称", "本次提供数据字段", "网址"], 1):
        c = sw.cell(3, ci, h)
        c.font = Font(name="微软雅黑", bold=True, size=9)
        c.fill = shfill; c.border = tb; c.alignment = ha
    sw.row_dimensions[3].height = 22

    for i, (cat, nm, fd, url) in enumerate(src_data, 4):
        sw.cell(i, 1, cat).font = df; sw.cell(i, 1).alignment = ha; sw.cell(i, 1).border = tb
        sw.cell(i, 2, nm).font = df; sw.cell(i, 2).alignment = al; sw.cell(i, 2).border = tb
        sw.cell(i, 3, fd).font = df; sw.cell(i, 3).alignment = al; sw.cell(i, 3).border = tb
        c = sw.cell(i, 4, url)
        c.font = Font(name="微软雅黑", size=9, color="0563C1", underline="single")
        c.hyperlink = url; c.alignment = al; c.border = tb
        sw.row_dimensions[i].height = 24

    sw.column_dimensions["A"].width = 18
    sw.column_dimensions["B"].width = 46
    sw.column_dimensions["C"].width = 48
    sw.column_dimensions["D"].width = 50

    sn = len(src_data) + 5
    remarks = [
        "备注:",
        "1. 每个数据单元格右上角红三角 = 该数据点的具体来源，同一行不同列来源不同，全程可追溯。",
        "2. 火灾危险性类别仅供参考，实际分类应以当地消防部门核定为准。",
        "3. 中国法规数值优先采用。",
        "4. 因牌号未定的物质，需查阅实际供应商SDS获取准确物性数据。",
    ]
    for i, r in enumerate(remarks):
        sw.cell(sn + i, 1, r).font = Font(name="微软雅黑", bold=(i == 0), size=9)

    # 安全写入
    tmp = os.path.join(output_dir, "chem_tmp.xlsx")
    wb.save(tmp)
    import zipfile
    zipfile.ZipFile(tmp).close()
    target = os.path.join(output_dir, f"{project_name}化学品物性数据.xlsx")
    if os.path.exists(target):
        os.remove(target)
    os.rename(tmp, target)
    print(f"OK: {target}")
    print(f"  {len(chemicals)} chemicals x {len(FIELDS)} fields")
    print(f"  {len(src_data)} sources in summary sheet")
    return target

if __name__ == "__main__":
    if not chemicals:
        print("请先在 chemicals 列表中填入化学品数据")
    else:
        build_excel(chemicals, PROJECT_NAME, OUTPUT_DIR, SRC_DATA)
